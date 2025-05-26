import openpyxl
from openpyxl.styles import PatternFill
from openpyxl.formatting.rule import FormulaRule
import os

file_path = "TEAM_FEATURES.xlsx"

likert_headers = [
    "I enjoy designing user experiences, visuals, or branding (Hipster – Designer).",
    "I like building and testing products or prototypes (Hacker – Builder).",
    "I enjoy pitching ideas, talking to customers, or persuading others (Hustler – Sales/Outreach).",
    "I prefer organizing people, tracking progress, and making things run smoothly (Handler – Operations/Project Management).",
    "I’m open to stepping into whatever role the team needs, even if it’s outside my comfort zone.",
    "I want to launch a real business through this course.",
    "I want to learn as much as I can about entrepreneurship.",
    "I primarily want to complete this capstone so I can finish my degree.",
    "I’m interested in learning how to use AI tools more effectively.",
    "I would like more structure or guidance on how to get value from AI tools.",
    "I prefer watching videos to learn new material.",
    "I prefer reading text-based materials such as articles or slides.",
    "I prefer listening to audio files or podcasts when learning.",
    "I prefer learning through interactive tools such as simulations, quizzes, or drag-and-drop activities.",
    "I prefer content that allows me to test my understanding as I go."
]

wb = openpyxl.load_workbook(file_path)
ws = wb.active

green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
yellow_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")

header_row = 1
header_to_col = {cell.value: cell.column_letter for cell in ws[header_row] if cell.value in likert_headers}

for header, col in header_to_col.items():
    ws.conditional_formatting.add(f"{col}2:{col}{ws.max_row}",
        FormulaRule(formula=[f'OR({col}2="Agree",{col}2="Strongly Agree")'], fill=green_fill))
    ws.conditional_formatting.add(f"{col}2:{col}{ws.max_row}",
        FormulaRule(formula=[f'OR({col}2="Disagree",{col}2="Strongly Disagree")'], fill=red_fill))
    ws.conditional_formatting.add(f"{col}2:{col}{ws.max_row}",
        FormulaRule(formula=[f'{col}2="Neutral"'], fill=yellow_fill))

wb.save("TEAM_FEATURES_COLORED.xlsx")
print("Conditional formatting applied and saved as TEAM_FEATURES_COLORED.xlsx")
